import os
import re
import time
import json
import requests
import psycopg2
from datetime import datetime, timedelta
import threading
import openpyxl
import logging
import aiohttp
from urllib.parse import quote

# Библиотека python-telegram-bot
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes

# ==============================
# Читаем переменные окружения (секретные данные)
# ==============================
TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "")
ADMIN_CHAT_ID = os.environ.get("ADMIN_CHAT_ID", "")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
YANDEX_DISK_TOKEN = os.environ.get("YANDEX_DISK_TOKEN", "")

# Параметры PostgreSQL
DATABASE_URL = os.environ.get("DATABASE_URL")

# Настройка логгера
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

#Переменная задержки ответа на сообщение

DELAY_SECONDS = 60

# ==============================
# Пути к файлам
# ==============================
knowledge_base_path = "knowledge_base.json"
prompt_path = "prompt.txt"
logs_directory = "dialog_logs"

# ==============================
# Прочитаем базу знаний и промпт
# ==============================
if not os.path.exists(logs_directory):
    os.makedirs(logs_directory, exist_ok=True)

with open(knowledge_base_path, "r", encoding="utf-8") as f:
    knowledge_base = json.load(f)

with open(prompt_path, "r", encoding="utf-8") as f:
    custom_prompt = f.read().strip()

# ==============================
# Сервисные переменные
# ==============================
gemini_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={GEMINI_API_KEY}"

# Для хранения истории диалогов (по user_id)
dialog_history_dict_tg = {}

# Для хранения (user_id -> (first_name, last_name)) и (user_id -> путь к лог-файлу)
user_names_tg = {}
user_log_files_tg = {}

# Лог-файл по умолчанию (используем, пока не знаем имени пользователя)
log_file_path = os.path.join(
    logs_directory,
    f"dialog_{datetime.utcnow().strftime('%Y-%m-%d_%H-%M-%S')}_unknown_user.txt"
)

# =========================================================================
# 1. ФУНКЦИЯ ПОИСКА ПОКУПОК КЛИЕНТОВ, ЕСЛИ В ЗАПРОСЕ ЕСТЬ ЕМЕЙЛ ИЛИ ТЕЛЕФОН
# =========================================================================

def get_client_info(user_question, user_id):
    """
    Анализирует запрос пользователя на наличие email и номера телефона.
    Если они найдены, ищет информацию о клиенте в excel файле clients.xlsx.
    Возвращает строку с информацией о клиенте или сообщение об отсутствии данных.
    Учитывает, что у клиента может быть несколько покупок.
    """

    client_info = ""

    # Регулярное выражение для поиска email
    email_regex = r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"
    emails = re.findall(email_regex, user_question)

    # Регулярное выражение для поиска номера телефона (учитываем разные варианты)
    phone_regex = r"(?:\+7|7|8)?[\s\-]?\(?(\d{3})\)?[\s\-]?(\d{3})[\s\-]?(\d{2})[\s\-]?(\d{2})"
    phones = re.findall(phone_regex, user_question)

    logging.info(f"Пользователь {user_id}: запросил информацию в таблице.")

    # Загружаем excel файл
    try:
        workbook = openpyxl.load_workbook("clients.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        logging.error("Файл clients.xlsx не найден.")
        return ""

    # Ищем информацию по email
    if emails:
        for email in emails:
            email_lower = email.lower().strip()
            logging.info(f"Пользователь {user_id}: ищем данные по емейлу {email_lower}.")
            email_found = False
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Email в столбце E (индекс 4, если A=0)
                cell_value = str(row[4]).lower().strip() if len(row) > 4 and row[4] else ""
                if email_lower == cell_value:
                    client_data = ", ".join(str(x) for x in row if x is not None)
                    client_info += f"Данные по клиенту, найденные по емейлу {email_lower}: {client_data}\n"
                    logging.info(f"Пользователь {user_id}: найдены данные по емейлу {email_lower}.")
                    email_found = True

            if not email_found:
                client_info += f"Данные по емейлу {email_lower} клиента не найдены\n"
                logging.warning(f"Пользователь {user_id}: не найдены данные по емейлу {email_lower}.")

    # Ищем информацию по номеру телефона
    if phones:
        for phone in phones:
            # Удаляем все, кроме цифр из найденного номера
            digits_only = "".join(filter(str.isdigit, "".join(phone)))
            logging.info(f"Пользователь {user_id}: ищем данные по телефону (только цифры): {digits_only}.")
            phone_found = False
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Телефон в столбце F (индекс 5)
                phone_cell = str(row[5]).strip() if len(row) > 5 and row[5] else ""
                # Убираем все нецифровые символы для надежности
                phone_digits = "".join(filter(str.isdigit, phone_cell))

                # Сравниваем последние 10 цифр
                if phone_digits and phone_digits[-10:] == digits_only[-10:]:
                    client_data = ", ".join(str(x) for x in row if x is not None)
                    client_info += f"Данные по клиенту, найденные по номеру телефона: {client_data}\n"
                    logging.info(f"Пользователь {user_id}: найдены данные по телефону {digits_only}.")
                    phone_found = True

            if not phone_found:
                client_info += f"Данные по телефону клиента не найдены\n"
                logging.warning(f"Пользователь {user_id}: не найдены данные по телефону.")

    return client_info

# ==============================
# 2. ФУНКЦИЯ УВЕДОМЛЕНИЙ
# ==============================

def send_operator_notification(dialog_id, initial_question, dialog_summary, reason_guess, first_name="", last_name=""):
    """
    Уведомление, если пользователь запросил оператора в процессе диалога.
    """
    message = f"""
🆘 Запрос оператора!
👤 Пользователь: {first_name} {last_name}
Изначальный вопрос клиента: {initial_question}
Сводка обсуждения: {dialog_summary}
Предполагаемая причина: {reason_guess}
    """.strip()

    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    data = {
        "chat_id": ADMIN_CHAT_ID,
        "text": message,
        "parse_mode": "Markdown"
    }
    requests.post(url, data=data)


# ==============================
# 3. РАБОТА С ЯНДЕКС.ДИСКОМ: ЗАГРУЗКА ЛОГ-ФАЙЛОВ
# ==============================
def upload_log_to_yandex_disk(log_file_path):
    # Проверяем, существует ли папка на Яндекс.Диске
    create_dir_url = "https://cloud-api.yandex.net/v1/disk/resources"
    headers = {"Authorization": f"OAuth {YANDEX_DISK_TOKEN}"}
    params = {"path": "disk:/app-logs"}
    requests.put(create_dir_url, headers=headers, params=params)

    if not os.path.exists(log_file_path):
        return

    if not YANDEX_DISK_TOKEN:
        print("YANDEX_DISK_TOKEN не задан. Пропускаем загрузку логов.")
        return

    file_name = os.path.basename(log_file_path)
    ya_path = f"disk:/app-logs/{file_name}"

    get_url = "https://cloud-api.yandex.net/v1/disk/resources/upload"
    params = {
        "path": ya_path,
        "overwrite": "true"
    }
    headers = {
        "Authorization": f"OAuth {YANDEX_DISK_TOKEN}"
    }
    r = requests.get(get_url, headers=headers, params=params)
    if r.status_code != 200:
        print("Ошибка при получении URL для загрузки на Яндекс.Диск:", r.text)
        return

    href = r.json().get("href", "")
    if not href:
        print("Не нашли 'href' в ответе Яндекс.Диска:", r.text)
        return

    with open(log_file_path, "rb") as f:
        upload_resp = requests.put(href, files={"file": f})
        if upload_resp.status_code == 201:
            print(f"Лог-файл {file_name} успешно загружен на Яндекс.Диск.")
        else:
            print("Ошибка загрузки на Яндекс.Диск:", upload_resp.text)


# ==============================
# 4. СОХРАНЕНИЕ ДИАЛОГОВ В POSTGRES
# ==============================
def store_dialog_in_db(user_id, user_message, bot_message, client_info=""):
    """
    Сохраняем каждую пару (user_message + bot_message) в базу PostgreSQL.
    Добавлено сохранение информации о клиенте.
    """
    try:
        conn = psycopg2.connect(DATABASE_URL)
        cur = conn.cursor()

        # Создание таблицы, если её нет
        cur.execute("""
            CREATE TABLE IF NOT EXISTS dialogues (
                id SERIAL PRIMARY KEY,
                user_id BIGINT,
                user_message TEXT,
                bot_message TEXT,
                client_info TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Вставка записи
        cur.execute(
            """INSERT INTO dialogues (user_id, user_message, bot_message, client_info)
                VALUES (%s, %s, %s, %s)""",
            (user_id, user_message, bot_message, client_info)
        )

        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        logging.error(f"Ошибка при сохранении диалога в БД: {e}")


def load_dialog_from_db(user_id):
    """
    Подгрузить из БД всю историю сообщений для указанного user_id.
    Возвращает список словарей вида: [{"user": "...", "bot": "...", "client_info": "..."}, ...].
    """
    dialog_history = []
    try:
        conn = psycopg2.connect(DATABASE_URL)
        cur = conn.cursor()
        # Берём все сообщения по user_id, сортируем по id (или по created_at)
        cur.execute("""
            SELECT user_message, bot_message, client_info
            FROM dialogues
            WHERE user_id = %s
            ORDER BY id ASC
        """, (user_id,))

        rows = cur.fetchall()
        for row in rows:
            user_m = row[0]
            bot_m = row[1]
            client_info = row[2]
            dialog_history.append({"user": user_m, "bot": bot_m, "client_info": client_info})

        cur.close()
        conn.close()
    except Exception as e:
        print(f"Ошибка при загрузке диалога из БД для user_id={user_id}:", e)

    return dialog_history


# ==============================
# 5. ЛОГИРОВАНИЕ
# ==============================
def log_dialog(user_question, bot_response, relevant_titles, relevant_answers, user_id, full_name="", client_info=""):
    """Логируем в локальный файл + отправляем пару (user_message, bot_message) в PostgreSQL.
    Без подсчёта токенов.
    """
    # Сохраняем в базу данных
    store_dialog_in_db(user_id, user_question, bot_response, client_info)

    current_time = datetime.utcnow() + timedelta(hours=6)
    formatted_time = current_time.strftime("%Y-%m-%d_%H-%M-%S")

    # Определяем лог-файл для пользователя
    if user_id in user_log_files_tg:
        local_log_file = user_log_files_tg[user_id]
    else:
        local_log_file = log_file_path

    # Пишем данные в лог-файл
    with open(local_log_file, "a", encoding="utf-8") as log_file:
        log_file.write(f"[{formatted_time}] {full_name}: {user_question}\n")
        if relevant_titles and relevant_answers:
            for title, answer in zip(relevant_titles, relevant_answers):
                log_file.write(f"[{formatted_time}] Найдено в базе знаний: {title} -> {answer}\n")
        if client_info:
            log_file.write(f"[{formatted_time}] Информация по клиенту: {client_info}\n")
        log_file.write(f"[{formatted_time}] Модель: {bot_response}\n\n")

    print(f"Содержимое лога:\n{open(local_log_file, 'r', encoding='utf-8').read()}")

    # Загружаем лог-файл в Яндекс.Диск
    upload_log_to_yandex_disk(local_log_file)


# ==============================
# 6. ИНТЕГРАЦИЯ С GEMINI
# ==============================
def find_relevant_titles_with_gemini(user_question):
    titles = list(knowledge_base.keys())
    prompt_text = f"""
Вот список вопросов-ключей:
{', '.join(titles)}

Найди три наиболее релевантных вопроса к запросу: "{user_question}".
Верни только сами вопросы, без пояснений и изменений.
    """.strip()

    data = {
        "contents": [{"parts": [{"text": prompt_text}]}]
    }
    headers = {"Content-Type": "application/json"}

    for attempt in range(5):
        resp = requests.post(gemini_url, headers=headers, json=data)
        if resp.status_code == 200:
            try:
                result = resp.json()
                text_raw = result['candidates'][0]['content']['parts'][0]['text']
                lines = text_raw.strip().split("\n")
                return [ln.strip() for ln in lines if ln.strip()]
            except KeyError:
                return "Извините, произошла ошибка при обработке ответа модели."
        elif resp.status_code == 500:
            time.sleep(10)
        else:
            return f"Ошибка: {resp.status_code}. {resp.text}"
    return "Извините, я сейчас не могу ответить. Попробуйте позже."


def generate_response(user_question, client_data, dialog_history, custom_prompt, relevant_answers=None):
    # Формируем историю в текстовом виде
    history_text = "\n".join([
        f"Пользователь: {turn.get('user','Неизвестно')}\nМодель: {turn.get('bot','Нет ответа')}"
        for turn in dialog_history
    ])

    knowledge_hint = (
        f"Подсказки из базы знаний: {relevant_answers}"
        if relevant_answers else ""
    )

    full_prompt = (
        f"{custom_prompt}\n\n"
f"Контекст диалога:\n{history_text}\n\n"
        f"Пользователь: {user_question}\n"
        f"Информация о клиенте: {client_data}\n"
        f"Модель:"
    )

    data = {
        "contents": [{"parts": [{"text": full_prompt}]}]
    }
    headers = {"Content-Type": "application/json"}

    for attempt in range(5):
        resp = requests.post(gemini_url, headers=headers, json=data)
        if resp.status_code == 200:
            try:
                result = resp.json()
                return result['candidates'][0]['content']['parts'][0]['text']
            except KeyError:
                return "Извините, произошла ошибка при обработке ответа модели."
        elif resp.status_code == 500:
            time.sleep(10)
        else:
            return f"Ошибка: {resp.status_code}. {resp.text}"
    return "Извините, я сейчас не могу ответить. Попробуйте позже."


def generate_summary_and_reason(dialog_history):
    history_text = " | ".join([
        f"Пользователь: {turn.get('user', 'Неизвестно')} -> Модель: {turn.get('bot', 'Нет ответа')}"
        for turn in dialog_history[-10:]
    ])
    prompt_text = f"""
Сводка диалога: {history_text}

На основе предоставленного диалога:
1. Сформируй сводку обсуждения.
2. Предположи причину, почему пользователь запросил оператора.
    """.strip()
    data = {
        "contents": [{"parts": [{"text": prompt_text}]}]
    }
    headers = {"Content-Type": "application/json"}

    for attempt in range(5):
        resp = requests.post(gemini_url, headers=headers, json=data)
        if resp.status_code == 200:
            try:
                result = resp.json()
                output = result['candidates'][0]['content']['parts'][0]['text'].split("\n", 1)
                dialog_summary = output[0].strip() if len(output) > 0 else "Сводка не сформирована"
                reason_guess = output[1].strip() if len(output) > 1 else "Причина не определена"
                return dialog_summary, reason_guess
            except (KeyError, IndexError):
                return "Сводка не сформирована", "Причина не определена"
        elif resp.status_code == 500:
            time.sleep(10)
        else:
            return "Ошибка API", "Ошибка API"
    return "Не удалось связаться с сервисом", "Не удалось связаться с сервисом"


# ==============================
# 7. ТАЙМЕР, БУФЕР СООБЩЕНИЙ и ОБРАБОТЧИК СООБЩЕНИЙ
# ==============================
user_buffers_tg = {}  # Буферы сообщений для каждого пользователя
user_timers_tg = {} # Таймеры для каждого пользователя
last_questions_tg = {} # Последний вопрос каждого пользователя (для логики нажатия на "оператора")

# ==============================
# 8. ПАУЗА ДЛЯ КОНКРЕТНОГО ПОЛЬЗОВАТЕЛЯ
# ==============================

def is_user_paused(user_id):
    """
    Проверяет, находится ли пользователь на паузе, отправляя запрос в бот уведомлений.
    """
    try:
        # Определение платформы
        if isinstance(user_id, int):
            platform = "tg"
            identifier = str(user_id)
        elif isinstance(user_id, str):
            platform = "vk"
            identifier = user_id
        else:
            logging.error(f"Неподдерживаемый тип идентификатора пользователя: {type(user_id)}")
            return False

        url = f"http://telegram-bot.railway.internal.com/is_paused/{platform}/{quote(identifier)}" # Изменено
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            data = response.json()
            return data.get("paused", False)
        else:
            logging.error(f"Ошибка проверки паузы для {identifier} на платформе {platform}: {response.status}")
            return False
    except Exception as e:
        logging.error(f"Ошибка подключения к боту уведомлений: {e}")
        return False

# ===================================================================
# Обработчики для Телеграма
# ===================================================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отправляет приветственное сообщение, когда пользователь запускает бота."""
    user_id = update.message.from_user.id
    first_name = update.message.from_user.first_name
    last_name = update.message.from_user.last_name
    full_name = f"{first_name} {last_name}".strip()

    # Сохраняем информацию о пользователе
    user_names_tg[user_id] = (first_name, last_name if last_name else "")

    # Формируем отдельный log_file_path c именем/фамилией
    now_str = datetime.utcnow().strftime('%Y-%m-%d_%H-%M-%S')
    custom_file_name = f"dialog_{now_str}_{full_name}.txt"
    custom_log_path = os.path.join(logs_directory, custom_file_name)
    user_log_files_tg[user_id] = custom_log_path

    # Загрузка истории диалога из базы данных
    dialog_history_dict_tg[user_id] = load_dialog_from_db(user_id)

    await context.bot.send_message(chat_id=update.effective_chat.id, text=f"Здравствуйте, {full_name}! Я Ваш персональный ассистент.")

async def handle_telegram_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает входящие текстовые сообщения от пользователя."""
    user_id = update.message.from_user.id
    text = update.message.text

    # Добавляем сообщение в буфер
    if user_id not in user_buffers_tg:
        user_buffers_tg[user_id] = []
    user_buffers_tg[user_id].append(text)

    # Сбрасываем/перезапускаем таймер
    if user_id in user_timers_tg:
        user_timers_tg[user_id].cancel()
    timer = threading.Timer(DELAY_SECONDS, generate_and_send_response_tg, args=(update, context))
    user_timers_tg[user_id] = timer
    timer.start()

async def generate_and_send_response_tg(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Генерирует ответ на основе запроса пользователя и отправляет его."""
    user_id = update.message.from_user.id

     # Проверяем, находится ли пользователь на паузе перед генерацией ответа
    first_name, last_name = user_names_tg.get(user_id, ("Неизвестно", ""))
    full_name = f"{first_name} {last_name}".strip()
    if is_user_paused(user_id): # Убрано await
        print(f"Пользователь {full_name} находится на паузе. Пропускаем генерацию ответа.")
        return

    # Проверяем, не инициирована ли остановка бота
    if user_id in user_timers_tg:
        user_timers_tg[user_id].cancel()
        del user_timers_tg[user_id]

    # Получаем полный текст сообщений из буфера
    combined_text = "\n".join(user_buffers_tg.get(user_id, []))
    if not combined_text:
        return
    user_buffers_tg[user_id] = [] # Очищаем буфер

    # Загружаем историю диалога из базы данных
    dialog_history = dialog_history_dict_tg.get(user_id, [])

    # Проверяем наличие email или телефона в сообщении
    emails = re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", combined_text)
    phones = re.findall(r"(?:\+7|7|8)?[\s\-]?\(?(\d{3})\)?[\s\-]?(\d{3})[\s\-]?(\d{2})[\s\-]?(\d{2})", combined_text)

    client_data = ""
    if emails or phones:
        # Ищем информацию о клиенте
        client_data = get_client_info(combined_text, user_id)
        logging.info(f"Пользователь {user_id}: запрошена информация о клиенте из таблицы.")
        if client_data:
            logging.info(f"Пользователь {user_id}: найдена информация о клиенте: {client_data}")
        else:
            logging.info(f"Пользователь {user_id}: информация о клиенте не найдена.")

    # Ищем релевантные ответы в базе знаний
    relevant_titles = find_relevant_titles_with_gemini(combined_text)
    relevant_answers = [knowledge_base[t] for t in relevant_titles if t in knowledge_base]

    # Генерируем ответ от модели
    model_response = generate_response(combined_text, client_data, dialog_history, custom_prompt, relevant_answers)

    # Обновляем историю диалога
    dialog_history.append({"user": combined_text, "bot": model_response, "client_info": client_data})
    dialog_history_dict_tg[user_id] = dialog_history

    # Логируем диалог
    log_dialog(combined_text, model_response, relevant_titles, relevant_answers, user_id, full_name, client_data)

    # Проверяем, запрашивал ли пользователь оператора
    if "оператор" in combined_text.lower():
        summary, reason = generate_summary_and_reason(dialog_history)
        send_operator_notification(user_id, combined_text, summary, reason, first_name, last_name)

    # Отправляем ответ пользователю
    await context.bot.send_message(chat_id=update.effective_chat.id, text=model_response)

# ==============================
# 9. ГЛАВНАЯ ФУНКЦИЯ
# ==============================

def main():
    """Запускает бота."""
    try:
        application = ApplicationBuilder().token(TELEGRAM_TOKEN).build()

        start_handler = CommandHandler('start', start)
        message_handler = MessageHandler(filters.TEXT & (~filters.COMMAND), handle_telegram_message)

        application.add_handler(start_handler)
        application.add_handler(message_handler)

        print("Бот запущен и готов к работе.")

        application.run_polling()

    except Exception as e:
        logging.error(f"Произошла ошибка при запуске бота: {e}")

if __name__ == "__main__":
    main()